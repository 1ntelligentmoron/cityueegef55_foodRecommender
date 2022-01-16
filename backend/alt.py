# Original matrix factorisation system, based on:
# https://towardsdatascience.com/recommendation-system-matrix-factorization-d61978660b4b
# But is very slow and the factorisation had a low accuracy.


# Importing required modules
import numpy as np
import openpyxl


# Read Excel and output as NumPy array
def reviews():

    wb = openpyxl.load_workbook('C:\\Users\\Hayden\\Desktop\\ratings.xlsx')
    ws = wb.active

    reviews = np.zeros((9018, 1001), dtype=np.float64)
    for j in range(2, 81568):
        cell_col = f'A{j}'
        cell_row = f'B{j}'
        cell_val = f'C{j}'
        col = ws[cell_col].value - 1
        row = ws[cell_row].value - 1
        val = float(ws[cell_val].value)
        reviews[row][col] = val

    row_sums = np.sum(reviews, axis=1)
    fakes = []
    for k in range(9018):
        if row_sums[k] == 0:
            fakes.append(k)
    reviews = np.delete(reviews, fakes, axis=0)

    wb.close()
    return reviews


# Wrapper function for matrix factorisation
def mat_fact(mat_in, lfs: int, try_for):
    
    """
    mat_in: sparse matrix to be factorised
    lfs: number of latent features
    try_for: number of attempts"""
    
    # Hyperparameters
    a = 1 / try_for
    b = 0.02
    
    try:
        
        # Initialising 2 random matrices to be factorised into
        mat_col = np.random.rand(len(mat_in), lfs)  # Number of columns
        mat_row = np.random.rand(lfs, len(mat_in[0]))  # Number of rows

        # for step in pb.progressbar(range(try_for)):
        for step in range(try_for):

            for i in range(len(mat_in)):
                for j in range(len(mat_in[i])):
                    if mat_in[i][j] > 0:
                        # Error calculation
                        error = round(mat_in[i][j] - np.dot(mat_col[i,:],mat_row[:,j]), 2)
                        for k in range(lfs):
                            # Gradient calculation with hyperparameters
                            mat_col[i][k] = round(mat_col[i][k] + a * (2 * error * mat_row[k][j] - b * mat_col[i][k]), 2)
                            mat_row[k][j] = round(mat_row[k][j] + a * (2 * error * mat_col[i][k] - b * mat_row[k][j]), 2)

            # e_mat_in = np.dot(mat_col, mat_row)

            e = 0
            for i in range(len(mat_in)):
                for j in range(len(mat_in[i])):
                    if mat_in[i][j] > 0:
                        e += round(pow(mat_in[i][j] - np.dot(mat_col[i, :], mat_row[:, j]), 2), 2)
                        for k in range(lfs):
                            e += round((b/2) * (pow(mat_col[i][k], 2) + pow(mat_row[k][j], 2)), 2)

            if e < 0.05:
                return mat_col, mat_row.T, e  # Return matrix factors once error is acceptable
            
            if np.isnan(e):
                raise OverflowError
            
            print(e)  # Update progress
            
        return mat_col, mat_row.T, e  # Return anyways
    
    except KeyboardInterrupt:
        print('Quitted early: raised KeyboardInterrupt')
        return mat_col, mat_row.T, e  # (backend control) Return immediately


# Script
if __name__ == '__main__':
    # halter = input('Halted (press RETURN to continue)')
    mat_reviews = reviews()
    mat_restaurant_lf, mat_user_lf, devi = mat_fact(mat_reviews, 10, 500)
    filled = np.dot(mat_restaurant_lf, mat_user_lf.T)
    np.savetxt('../data/ranked.csv', filled, delimiter=',')
