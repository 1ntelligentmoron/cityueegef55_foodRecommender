from backend import cf, r
import pandas as pd
import openpyxl
from random import randint as rnd
wb = openpyxl.load_workbook('backend/r.xlsx')
ws = wb.active


def recommend(budget, range, lat, long):
    
    """
    Three cases:
    Case 1: user is new (indentified by the lack of 'r_filled.csv' in /backend/)
    Case 2: user no longer a cold start problem (indicated by cell ALZ6 in 'r.xlsx')
    Case 3: user's preference matrix (r_filled.csv) has been created
    """
    
    # Test for case 1
    try:
        
        # Case 3
        df = pd.read_csv('backend/r_filled.csv', header=None)  # Raises FileNotFoundError if preference matrix not generated
        df = df.iloc[:, -1:]  # Extracting column storing user preference
        
        recs = []
        while len(recs) < 3:
            max_pos = df.idxmax().to_list()
            df[1000][max_pos] = 0
            max_id = int(max_pos[0]) + 1
            print('Finding {}/3 match(es), id={}'.format(len(recs)+1, max_id))
            if r.check(max_id, budget, range, lat, long):
                recs.append(max_id)
                print('Found {}/3 match(es)'.format(len(recs)))
    
    except FileNotFoundError:
        
        if not not ws['ALZ6'].value:  # not not x is faster than bool(x)
            # Case 2
            try:
                return 'CF_NOW'
            finally:
                cf.main()
            
        else:
            # Case 1
            recs = []
            while len(recs) < 3:
                id = rnd(1, 5167)
                if r.check(id, budget, range, lat, long):
                    recs.append(id)
    
    r1, kw1, addr1, coord1, r_id1 = r.info(recs[0], True)
    r2, kw2, addr2, coord2, r_id2 = r.info(recs[1], True)
    r3, kw3, addr3, coord3, r_id3 = r.info(recs[2], True)
    
    return ((r1, kw1, addr1, coord1, r_id1), (r2, kw2, addr2, coord2, r_id2), (r3, kw3, addr3, coord3, r_id3))


def chosen(id):
    
    info = r.info(id)
    
    # Do nothing if preference matrix already generated
    try:
        with open('backend/r_filled.csv', 'r') as f:
            pass
    
    except FileNotFoundError:
        
        row = id + 1
        
        # Increase rating of chosen restaurant by 1.0, limited to 5.0
        rating_new_chosen = round(float(ws[f'ALU{row}'].value) + 1.0, 3)
        if rating_new_chosen > 5.0:
            rating_new_chosen = 5.0
        ws[f'ALU{row}'] = rating_new_chosen
        
        # Increase rating of related restaurant(s) by 0.5, limited to 4.0
        for i in range(1, 5169):
            if i == row:
                continue  # Do not add 0.5 to chosen restaurant
            kw1 = ws[f'D{i}'].value
            kw2 = ws[f'E{i}'].value
            
            kws_req = info['keywords']
            if kw1 in kws_req and kw2 in kws_req:  # If restaurant is related to chosen restaurant
                rating_new_related = round(float(ws[f'ALU{i}'].value) + 0.5, 3)
                if rating_new_related < 4.5:
                    ws[f'ALU{row}'] = rating_new_related
    
    return 'You have chosen {} (Address: {}).'.format(info['name'], info['address'])
