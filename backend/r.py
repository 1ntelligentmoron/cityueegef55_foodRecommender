import openpyxl
from geopy import distance
wb = openpyxl.load_workbook('backend/r.xlsx', read_only=True)
ws = wb.active


def info(id, as_output=False):
    
    info = {
        'name': '',  # column B
        'b_range': -1,  # column C
        'keywords': (None, None,),  # columns D, E
        'address': '',  # column F
        'coords': (None, None,),  # columns G, H
    }
    
    row = id + 1
    
    name = ws[f'B{row}'].value
    info['name'] = f'{name}'
    
    info['b_range'] = int(ws[f'C{row}'].value)
    
    broad = ws[f'D{row}'].value
    narrow = ws[f'E{row}'].value
    info['keywords'] = (f'{broad}', f'{narrow}',)
    
    address = ws[f'F{row}'].value
    info['address'] = f'{address}'
    
    lat = ws[f'G{row}'].value
    long = ws[f'H{row}'].value
    info['coords'] = (float(lat), float(long),)
    
    if as_output:
        return info['name'], info['keywords'], info['address'], info['coords'], id
    return info


def check(id, u_budget, u_range, u_lat, u_long):
    
    # Lookup budget range of user (referenced with r.xlsx)
    if 20 <= u_budget < 50:
        u_brange = 1
    elif 50 <= u_budget < 100:
        u_brange = 2
    elif 100 <= u_budget < 200:
        u_brange = 3
    elif 200 <= u_budget < 400:
        u_brange = 4
    elif 400 <= u_budget < 800:
        u_brange = 5
    else:
        u_brange = 6
    
    ok_budget = info(id)['b_range'] <= u_brange  # Check budget
    ok_range = round(distance.distance((u_lat, u_long), info(id)['coords']).km, 1) <= u_range  # Check in range
    print('Budget OK: {}\nRange OK: {}'.format(ok_budget, ok_range))
    
    return ok_budget and ok_range
